#
# Title: Read data from .xlsx file/ .csv file  in Python an performs regression
# Date: 06 October 2021
# Description: This program that read or exract the data from .xlxs file/ .csv file in Python and reflect the output in command prompt or IDLE 
#
# Copyright: Ruchi Sharma Â© 2021
# 
# @author  Ruchi Sharma
 
# @version 1.0.1	2021-10-06  Perform regression using Linear equation and cofficient regression method. 
# 

#import all the required library
from tkinter import * 
from tkinter import messagebox
from tkinter import filedialog as fd
from tkinter import ttk
from scipy import stats
from tkinter.messagebox import showinfo
from scipy.stats import spearmanr
from scipy import optimize
import math
import xlrd 
import xlwt
import csv
import openpyxl
import matplotlib.pyplot as plt 
import numpy as np
import tkinter as tk

#declaring gui with size and title 
top = Tk()  
top.geometry("600x700")
top.title("GUI APP - Regression in Python")

# workbook is created
#wb= Workbook()

#this sring is to show the user the output in the GUI
output = StringVar()

#these arrays stores the input data fron excel sheet
Data=[]
X=[]
Y=[]

#this routine calls the browser option and waits for the uder to select an xlsx file. Then it 
# gets the datas from that sheets and stored them in arrays. After getting the data in array it
# puts the data in the tree and shows it in the GUI
def brows():
	xlrd.xlsx.ensure_elementtree_imported(False, None)
	xlrd.xlsx.Element_has_iter = True
	name=fd.askopenfilename()
	loc=(name)
	wb = xlrd.open_workbook(loc)
	worksheet = wb.sheet_by_name('Sheet1') 
	worksheet = wb.sheet_by_index(0) 	
	i=1
	while (i<11):
		x = worksheet.cell(i, 0)
		y = worksheet.cell(i, 1)
		print(" ", int(x.value), " ", int(y.value)) 		
		Data.append((int(x.value),int(y.value)))
		X.append(int(x.value))
		Y.append(int(y.value))
		i=i+1
	for a in Data:
		tree.insert('', tk.END, values=a)
	def item_selected(event):
		for selected_item in tree.selection():
			item = tree.item(selected_item)
			record = item['values']
			showinfo(title='Information',message=','.join(record))
	tree.bind('<<TreeviewSelect>>', item_selected)



#This routine uses the datas from the xlsx file to calculate Linear regression
def coeff():	
	Exy=Ex=Ey=Ex2=Ey2=i=0
	while(i<10):
		Ex += X[i]
		Ey += Y[i]
		Ex2 += (X[i]*X[i])
		Ey2 += (Y[i]*Y[i])
		Exy += (X[i]*Y[i])
		i=i+1
	nExy= Exy*10
	nEx2= Ex2*10
	ExEy= Ex*Ey
	nEy2 = Ey2*10
	Ex2 = Ex*Ex
	Ey2 = Ey*Ey
	bxy=(nExy-Ex*Ey)/(nEy2-Ey2)
	bxy=round(bxy,2-int(math.floor(math.log10(abs(bxy))))-1)
	byx=(nExy-Ex*Ey)/(nEx2-Ex2)
	byx=round(byx,2-int(math.floor(math.log10(abs(byx))))-1)
	print("Regression coefficient for X on Y: ",bxy)
	print("Regression coefficient for X on Y: ",byx)
	strr="bXY: "+str(bxy)+"\nbYX: "+str(byx)
	output.set(strr)
	write= ['Correlation']
	filename = "Regression_Coefficient.csv"
	r= [[bxy]]	
	# writing to csv file
	with open(filename, 'w') as csvfile:
		csvwriter = csv.writer(csvfile)
		csvwriter.writerow(write)
		csvwriter.writerows(r)

#this routine uses the datas from the xlsx file to calculate coorelation using spearman rank's  correlation method
def eqq():
	Exy=Ex=Ey=Ex2=Ey2=i=0
	while(i<10):
		Ex += X[i]
		Ey += Y[i]
		Ex2 += (X[i]*X[i])
		Ey2 += (Y[i]*Y[i])
		Exy += (X[i]*Y[i])
		i=i+1
	nExy= Exy*10
	nEx2= Ex2*10
	ExEy= Ex*Ey
	nEy2 = Ey2*10
	Ex2 = Ex*Ex
	Ey2 = Ey*Ey
	bxy=(nExy-Ex*Ey)/(nEy2-Ey2)
	bxy=round(bxy,2-int(math.floor(math.log10(abs(bxy))))-1)
	byx=(nExy-Ex*Ey)/(nEx2-Ex2)
	byx=round(byx,2-int(math.floor(math.log10(abs(byx))))-1)
	print("Regression coefficient for X on Y: ",bxy)
	print("Regression coefficient for X on Y: ",byx)
	strr="Y-"+str(byx)+"X="+ str(Ey-byx*Ex)+"\nX-"+str(bxy)+"Y="+ str(Ex-bxy*Ey)
	
	#eqX.config(text="Regression equation are: "+"\n"+"Y-"+str(Ey)+ "="+str(bxy)
	#print("Regression equation are: "+"\n"+"Y-"+str(Ey)+"="+str(byx+"(X-"+ str(Ex)+")
	#eqX=coeff(y-Ey)+Ex
	#eqy =coeff(x-Ex)+Ey
	#z= eqX
	output.set(strr)
	#output.set("Pearson's Coorelation Coefficient: %.5f"%(byx))
	#output.set("Pearson's Coorelation Coefficient: %.5f"%(byx))
	
	write= ['Correlation']
	filename = "Regression Equations.csv"
	r= [[bxy]]	
	# writing to csv file
	with open(filename, 'w') as csvfile:
		csvwriter = csv.writer(csvfile)
		csvwriter.writerow(write)
		csvwriter.writerows(r)



#this routine graphically represents the datas from the xlsx file in a line graph 
def graph():
	
	plt.figure(figsize=(12,5))
	plt.scatter(X, Y)
	plt.grid()
	plt.title('Linear regression')
	plt.xlabel('X axis', fontsize=15)
	plt.ylabel('Y axis', fontsize=15)
	plt.plot(X, Y, solid_capstyle='round')
	#plt.scatter(x=x.mean(), y=y.mean(), marker='*', s=10**2.5, c='r') # average point
	plt.show()
	





#this is to say the user to select a file
heading= Label(top, text="Regression In Python",font=('Helvetica bold',20)).place(x=150,y=40)

header = Label(top, text="Select the .xlsx file to see the data:",font=('Helvetica bold',10)).place(x=40,y=80)

#this is to give the headings of the columns
cols= ('#1', '#2')

#this is for creating the tree
tree = ttk.Treeview(top, columns=cols, show='headings')

# this is for defining the column 
tree.heading('#1', text='X-Coordinates')
tree.heading('#2', text='Y-Coordinates')
tree.place(x=50,y=100,width=400,height=250) 

#this button helps user to browse for the xlsx file 
browse = Button(top, text = "Browse File", command=brows).place(x=230, y=370)

#this label will show the user the correlation of the datas
result = Label(top, text = "", textvariable=output, fg="green", font=('Helvetica bold',15)).place(x = 200,y = 420)
output.set("Result of the data will be shown here")

#these are the buttons which calls their repestive method functions to compute coorelation in that method 
coeff = Button(top, text = "Linear Regression Coefficient", command=coeff).place(x = 100, y = 500,width=160)
eqq = Button(top, text = "Linear Regression equation",command=eqq ).place(x = 300, y = 500,width=160)

graphBt = Button(top, text = "Graphical representation", command=graph).place(x = 200, y = 550,width=150)

#this button will close the UI
quit= Button(top, text = "QUIT", command = top.destroy, fg="red").place(x=240,y =650)

#runs the application
top.mainloop()